import os
import sys
import time
import datetime

import numpy as np
import pandas as pd
from pandas import Series, DataFrame
from openpyxl import load_workbook

#策略配置
LIMIT_PRICE = 5
LIMIT_AMOUNT = 30000
LIMIT_RANK_BETA = 0.0
LIMIT_RANK_MC = 0.0
LIMIT_RANK_PB = 0.0
LIMIT_N_IND = 21
LIMIT_N = 777
LIMIT_DAYS_LIST = 255
LIMIT_SUSPEND = True
LIMIT_BREAKER = True
WHITE_INDUSTRY = False
#字段名

#路径名
PROJECT_PATH = 'D:/stock'
OUTPUT_PATH = PROJECT_PATH + '/Output'
# GLOBALCONFIG_PATH = PROJECT_PATH + '/Codes'
GLOBALCONFIG_PATH = PROJECT_PATH + '/Codes/Utility'
DATABASE_PATH = PROJECT_PATH + '/DataBase'
FACTORBASE_PATH = PROJECT_PATH + '/FactorBase'
LABELBASE_PATH = PROJECT_PATH + '/LabelBase'
SINGLEFACTOR_PATH = PROJECT_PATH + '/SingleFactor'
MULTIFACTOR_PATH = PROJECT_PATH + '/MultiFactor'
MODEL_PATH = PROJECT_PATH + '/Model'
IC_PATH = PROJECT_PATH + '/IC'
BACKTEST_PATH = PROJECT_PATH + '/Backtest'
PREDICT_PATH = PROJECT_PATH + '/Predict'

#财报时间
END_DATE = ['0331', '0630', '0930', '1231']
ANN_DATE = ['0430', '0831', '1031', '0430']
UPD_DATE = ['0507', '0907', '1107', '0507']
DELTA_DATE = [50, 70, 40, 130]

#风险因子
WHITE_INDUSTRY_DIC = {}
excel = load_workbook('%s/行业.xlsx'%GLOBALCONFIG_PATH)
sheet_name = excel.sheetnames[0]  
sheet = excel[sheet_name]
n = 0
for line in sheet.iter_rows():
    if n == 0:
        pass
    n += 1
    if line[0].value not in WHITE_INDUSTRY_DIC.keys():
        WHITE_INDUSTRY_DIC[line[0].value] = {}
    if line[1].value not in WHITE_INDUSTRY_DIC[line[0].value].keys():
        WHITE_INDUSTRY_DIC[line[0].value][line[1].value] = []
    if line[2].fill.fgColor.rgb != 'FFFFFF00':
    # if True:
        WHITE_INDUSTRY_DIC[line[0].value][line[1].value].append(line[2].value)
WHITE_INDUSTRY_LIST = []
for k1 in WHITE_INDUSTRY_DIC.keys():
    for k2 in WHITE_INDUSTRY_DIC[k1].keys():
        WHITE_INDUSTRY_LIST.extend(WHITE_INDUSTRY_DIC[k1][k2])
